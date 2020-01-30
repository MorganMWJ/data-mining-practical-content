from openpyxl import load_workbook
wb = load_workbook( filename = 'prac1_datamining.xlsx')

import json

import pymongo
from pymongo import MongoClient
client = MongoClient('mongodb://mwj7:ZcnrlM19FX@nosql.dcs.aber.ac.uk/mwj7')
db = client.mwj7

#the active worksheet
ws = wb.active

people = []
#iterate over rows in spreadsheet
for row in ws.iter_rows(min_row=2):
	person = {}
	for cell in row:
		if not cell.value is None:
			#get column name for key
			attr = ws.cell(column=cell.col_idx,row=1).value
			#get cell value for value
			person[attr] = cell.value
	#insert json representation of person into mongo db here
	people.append(person)
	db.mwj7.insert(person)

print(str(people))
#json_output =  json.dumps(people)
#Todo: Future code could save the json to a file and then import that into mongo

